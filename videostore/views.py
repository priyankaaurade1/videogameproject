from django.utils import timezone
from django.shortcuts import render, redirect
from .models import *
import openpyxl
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
import random
from django.contrib import messages
from django.contrib.auth import authenticate, login
import pytz
from django.db.models import Max
from openpyxl.drawing.image import Image as ExcelImage
import os
from django.urls import reverse
from django.http import JsonResponse
from django.contrib.auth import logout
from django.contrib.auth.decorators import user_passes_test
import base64
from django.core.files.base import ContentFile
IST = pytz.timezone('Asia/Kolkata')

def custom_login(request):
    if request.user.is_authenticated:
        if request.user.role == 'superadmin':
            return redirect('superadmin_dashboard')
        elif request.user.role == 'staff':
            return redirect('customer_staff_entry')
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        if not username or not password:
            messages.error(request, "Please enter both username and password.")
            return render(request, 'login.html')

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            if user.role == 'superadmin':
                return redirect('superadmin_dashboard')
            elif user.role == 'staff':
                if user.store is None:
                    messages.error(request, "You are not assigned to any store. Contact admin.")
                    return redirect('custom_login')
                return redirect('customer_staff_entry')
            else:
                messages.error(request, "Invalid user role.")
        else:
            messages.error(request, "Invalid username or password.")
    return render(request, 'login.html')

def is_superadmin(user):
    return user.is_authenticated and user.is_superuser

@user_passes_test(is_superadmin, login_url='/forbidden/')
def superadmin_dashboard(request):
    if request.user.role != 'superadmin':
        messages.error(request, "Unauthorized access.")
        return redirect('custom_login')

    if request.method == 'POST':
        if 'new_store' in request.POST:
            name = request.POST.get('store_name')
            location = request.POST.get('store_location')
            if name:
                Store.objects.create(name=name, location=location)
                messages.success(request, "✅ New store added successfully!")
            else:
                messages.error(request, "⚠️ Store name is required.")
        
        if 'new_machine' in request.POST:
            name = request.POST.get('name')
            number = request.POST.get('number')
            store_id = request.POST.get('store_id')

            if name and number and store_id:
                try:
                    store = Store.objects.get(id=store_id)
                    Machine.objects.create(
                        name=name,
                        number=number,
                        store=store
                    )
                    messages.success(request, "✅ New machine added successfully!")
                except Store.DoesNotExist:
                    messages.error(request, "⚠️ Selected store does not exist.")
            else:
                messages.error(request, "⚠️ Machine name, number, and store are required.")

        if 'new_staff' in request.POST:
            username = request.POST.get('staff_username')
            email = request.POST.get('staff_email')
            password = request.POST.get('staff_password')
            store_id = request.POST.get('staff_store')
            phone_number = request.POST.get('staff_phone')
            if username and password:
                try:
                    store = Store.objects.get(id=store_id) if store_id else None
                    User.objects.create_user(
                        username=username,
                        email=email,
                        password=password,
                        role='staff',
                        store=store,
                        phone_number=phone_number,
                    )
                    messages.success(request, "✅ New staff user added successfully!")
                    request.session['clear_local_storage'] = True
                except Store.DoesNotExist:
                    messages.error(request, "⚠️ Selected store does not exist.")
            else:
                messages.error(request, "⚠️ Username and password are required for staff.")

        if 'edit_store' in request.POST:
            store_id = request.POST.get('store_id')
            name = request.POST.get('store_name')
            location = request.POST.get('store_location')
            try:
                store = Store.objects.get(id=store_id)
                store.name = name
                store.location = location
                store.save()
                messages.success(request, "✅ Store updated successfully!")
            except Store.DoesNotExist:
                messages.error(request, "⚠️ Store not found.")

        if 'edit_machine' in request.POST:
            machine_id = request.POST.get('machine_id')
            name = request.POST.get('name')
            number = request.POST.get('number')
            store_id = request.POST.get('store_id')
            try:
                machine = Machine.objects.get(id=machine_id)
                store = Store.objects.get(id=store_id)
                machine.name = name
                machine.number = number
                machine.store = store
                machine.save()
                messages.success(request, "✅ Machine updated successfully!")
            except (Machine.DoesNotExist, Store.DoesNotExist):
                messages.error(request, "⚠️ Machine or Store not found.")

        if 'edit_staff' in request.POST:
            staff_id = request.POST.get('staff_id')
            username = request.POST.get('staff_username')
            email = request.POST.get('staff_email')
            store_id = request.POST.get('staff_store')
            phone_number = request.POST.get('staff_phone')
            try:
                staff = User.objects.get(id=staff_id)
                staff.username = username
                staff.email = email
                staff.phone_number = phone_number
                staff.store = Store.objects.get(id=store_id) if store_id else None
                staff.save()
                messages.success(request, "✅ Staff updated successfully!")
            except (User.DoesNotExist, Store.DoesNotExist):
                messages.error(request, "⚠️ Staff or Store not found.")

        if 'delete_store' in request.POST:
            store_id = request.POST.get('store_id')
            try:
                store = Store.objects.get(id=store_id)
                store.delete()
                messages.success(request, "✅ Store deleted successfully!")
            except Store.DoesNotExist:
                messages.error(request, "⚠️ Store does not exist.")
        if 'delete_machine' in request.POST:
            machine_id = request.POST.get('machine_id')
            try:
                machine = Machine.objects.get(id=machine_id)
                machine.delete()
                messages.success(request, "✅ Machine deleted successfully!")
            except Machine.DoesNotExist:
                messages.error(request, "⚠️ Machine does not exist.")
        if 'delete_staff' in request.POST:
            staff_id = request.POST.get('staff_id')
            try:
                staff = User.objects.get(id=staff_id, role='staff')
                staff.delete()
                messages.success(request, "✅ Staff user deleted successfully!")
            except User.DoesNotExist:
                messages.error(request, "⚠️ Staff user does not exist.")

    stores = Store.objects.all()
    staff_users = User.objects.filter(role='staff')
    machines = Machine.objects.select_related('store').all()
    response = render(request, 'superadmin_dashboard.html', {
        'stores': stores,
        'staff_users': staff_users,
        'machines': machines,
    })
    if request.session.get('clear_local_storage'):
        del request.session['clear_local_storage']
    return response

def custom_logout(request):
    logout(request)
    return redirect('custom_login')

def generate_bill_no():
    india_tz = pytz.timezone("Asia/Kolkata")
    now = timezone.now().astimezone(india_tz)
    today_str = now.strftime("%d%m%y")  
    last_bill = GameData.objects.filter(bill_no__startswith=today_str).aggregate(Max('bill_no'))['bill_no__max']
    if last_bill:
        last_seq = int(last_bill[-3:])  
        next_seq = f"{last_seq + 1:03d}"
    else:
        next_seq = "001"

    return today_str + next_seq

# def export_report(request):
#     workbook = openpyxl.Workbook()
#     sheet = workbook.active
#     sheet.title = 'Game Report'
#     # Headers
#     headers = [
#         'Staff', 'Customer ID', 'Customer Name', 'Machine', 'In Points',
#         'Out Points', 'Good Luck', 'Expense Type','Expense Amount' 'Bill No', 'Date', 'Time', 
#     ]
#     sheet.append(headers)

#     for idx, entry in enumerate(GameData.objects.all(), start=2):
#         sheet.append([
#             entry.staff.username,
#             entry.customer_id,
#             entry.customer_name,
#             entry.machine,
#             entry.in_points,
#             entry.out_points,
#             entry.good_luck,
#             entry.expense_type,
#             entry.expense_amt,
#             entry.bill_no,
#             entry.date.strftime('%Y-%m-%d'),
#             entry.time.strftime('%H:%M:%S'),
#             "", 
#         ])

#     # Prepare response
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename=Staff Entries.xlsx'
#     workbook.save(response)
#     return response

@login_required
def export_staff_entries(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Staff Entries'

    # Headers
    headers = [
        'Staff', 'Customer ID', 'Customer Name', 'Machine', 'In Points',
        'Out Points', 'Good Luck', 'Expense Type', 'Expense Amount', 'Bill No', 'Date', 'Time'
    ]
    sheet.append(headers)

    # Filter entries
    if request.user.role == 'staff':
        queryset = GameData.objects.filter(entry_source='staff_entry', staff=request.user)
    else:
        queryset = GameData.objects.filter(entry_source='staff_entry')

    for entry in queryset.select_related('staff', 'machine__store'):
        machine_text = f"{entry.machine.name} - {entry.machine.number} ({entry.machine.store.name})" if entry.machine else "-"
        sheet.append([
            entry.staff.username if entry.staff else '-',
            entry.customer_id,
            entry.customer_name,
            machine_text,
            entry.in_points,
            entry.out_points,
            entry.good_luck,
            entry.expense_type,
            entry.expense_amt,
            entry.bill_no,
            entry.date.strftime('%Y-%m-%d'),
            entry.time.strftime('%H:%M:%S')
        ])

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Staff_Entries.xlsx'
    workbook.save(response)
    return response

@login_required
def export_staff_entries(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Staff Entries'

    # Headers
    headers = [
        'Staff', 'Customer ID', 'Customer Name', 'Machine', 'In Points',
        'Out Points', 'Good Luck', 'Expense Type', 'Expense Amount', 'Bill No', 'Date', 'Time'
    ]
    sheet.append(headers)

    # Filter entries
    if request.user.role == 'staff':
        queryset = GameData.objects.filter(entry_source='staff_entry', staff=request.user)
    else:
        queryset = GameData.objects.filter(entry_source='staff_entry')

    for entry in queryset.select_related('staff', 'machine__store'):
        machine_text = f"{entry.machine.name} - {entry.machine.number} ({entry.machine.store.name})" if entry.machine else "-"
        sheet.append([
            entry.staff.username if entry.staff else '-',
            entry.customer_id,
            entry.customer_name,
            machine_text,
            entry.in_points,
            entry.out_points,
            entry.good_luck,
            entry.expense_type,
            entry.expense_amt,
            entry.bill_no,
            entry.date.strftime('%Y-%m-%d'),
            entry.time.strftime('%H:%M:%S')
        ])

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Machine_Entries.xlsx'
    workbook.save(response)
    return response

@login_required
def export_customer_entries(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Customer Entries'

    headers = [
        'Staff', 'Customer ID', 'Customer Name', 'Machine', 'In Points',
        'Out Points', 'Good Luck', 'Expense Type', 'Expense Amount', 'Bill No', 'Date', 'Time'
    ]
    sheet.append(headers)

    # Filter entries
    if request.user.role == 'staff':
        queryset = GameData.objects.filter(entry_source='customer_staff_entry', staff=request.user)
    else:
        queryset = GameData.objects.filter(entry_source='customer_staff_entry')

    for entry in queryset.select_related('staff', 'machine__store'):
        machine_text = f"{entry.machine.name} - {entry.machine.number} ({entry.machine.store.name})" if entry.machine else "-"
        sheet.append([
            entry.staff.username if entry.staff else '-',
            entry.customer_id,
            entry.customer_name,
            machine_text,
            entry.in_points,
            entry.out_points,
            entry.good_luck,
            entry.expense_type,
            entry.expense_amt,
            entry.bill_no,
            entry.date.strftime('%Y-%m-%d'),
            entry.time.strftime('%H:%M:%S')
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Customer_Entries.xlsx'
    workbook.save(response)
    return response

@login_required
def staff_entry(request):
    if request.user.role == 'staff':
        if not request.user.store:
            messages.error(request, "You are not assigned to any store.")
            return redirect('custom_login')
        store = request.user.store
        machines = Machine.objects.filter(store=store)
        staff_machine = machines.first() if machines.count() == 1 else None
    elif request.user.is_superuser or request.user.role == 'superadmin':
        store = None
        machines = Machine.objects.all()
        staff_machine = None
    else:
        messages.error(request, "Unauthorized access.")
        return redirect('custom_login')
    now = timezone.now()

    if request.method == 'POST':
        entry_id = request.POST.get("entry_id")
        machine_id = request.POST.get('machine')
        photo_file = None
        photo_data = request.POST.get('photo_data')
        if photo_data:
            format, imgstr = photo_data.split(';base64,')
            ext = format.split('/')[-1]
            photo_file = ContentFile(base64.b64decode(imgstr), name=f"{random.randint(100000,999999)}.{ext}")

        if entry_id:
            # Editing existing record
            entry = GameData.objects.get(pk=entry_id)
            entry.customer_name = request.POST['customer_name']
            entry.machine = Machine.objects.get(pk=machine_id) if machine_id else None
            entry.in_points = request.POST['in_points']
            entry.out_points = request.POST['out_points']
            entry.good_luck = request.POST.get('good_luck') or 0
            entry.expense_type = request.POST['expense_type']
            entry.expense_amt = request.POST.get('expense_amt') or 0
            entry.remarks = request.POST['remarks']
            entry.date = now.date()
            entry.entry_source = 'staff_entry'
            # if photo_file:
            #     entry.photo = photo_file
            entry.save()
        else:
            # Creating new record
            entry = GameData.objects.create(
                staff=request.user,
                customer_id="cust-" + str(random.randint(10000, 99999)),
                customer_name=request.POST['customer_name'],
                machine=Machine.objects.get(pk=request.POST['machine']) if request.POST.get('machine') else None,
                in_points=request.POST['in_points'],
                out_points=request.POST['out_points'],
                good_luck=request.POST.get('good_luck') or 0,
                expense_type=request.POST['expense_type'],
                expense_amt=request.POST.get('expense_amt') or 0,
                bill_no=generate_bill_no(),
                # photo=photo_file,
                remarks=request.POST['remarks'],
                date=now.date(),
                time=now.time(),
                entry_source='staff_entry'
            )

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'id': entry.id})

        messages.success(request, "✅ Data saved successfully!")
        return redirect('staff_entry')

    context = {
        'now': now,
        'customer_id': "cust-" + str(random.randint(10000, 99999)),
        'bill_no': generate_bill_no(),
        'stores':store,
        'staff_machine':staff_machine,
        'machines':machines,
    }
    return render(request, 'staff_entry.html', context)

@login_required
def customer_staff_entry(request):
    if request.user.role == 'staff':
        if not request.user.store:
            messages.error(request, "You are not assigned to any store.")
            return redirect('custom_login')
        store = request.user.store
        machines = Machine.objects.filter(store=store)
        staff_machine = machines.first() if machines.count() == 1 else None
    elif request.user.is_superuser or request.user.role == 'superadmin':
        store = None
        machines = Machine.objects.all()
        staff_machine = None
    else:
        messages.error(request, "Unauthorized access.")
        return redirect('custom_login')
    now = timezone.now()

    if request.method == 'POST':
        entry_id = request.POST.get("entry_id")
        machine_id = request.POST.get('machine')
        photo_file = None
        photo_data = request.POST.get('photo_data')
        if photo_data:
            format, imgstr = photo_data.split(';base64,')
            ext = format.split('/')[-1]
            photo_file = ContentFile(base64.b64decode(imgstr), name=f"{random.randint(100000,999999)}.{ext}")

        if entry_id:
            # Editing existing record
            entry = GameData.objects.get(pk=entry_id)
            entry.customer_name = request.POST['customer_name']
            entry.machine = Machine.objects.get(pk=machine_id) if machine_id else None
            entry.in_points = request.POST['in_points']
            entry.out_points = request.POST['out_points']
            entry.good_luck = request.POST.get('good_luck') or 0
            entry.expense_type = request.POST['expense_type']
            entry.expense_amt = request.POST.get('expense_amt') or 0
            entry.remarks = request.POST['remarks']
            entry.date = now.date()
            entry.time = now.time()
            entry.entry_source = 'customer_staff_entry'
            if photo_file:
                entry.photo = photo_file
            entry.save()
        else:
            # Creating new record
            entry = GameData.objects.create(
                staff=request.user,
                customer_id="cust-" + str(random.randint(10000, 99999)),
                customer_name=request.POST['customer_name'],
                machine=Machine.objects.get(pk=request.POST['machine']) if request.POST.get('machine') else None,
                in_points=request.POST['in_points'],
                out_points=request.POST['out_points'],
                good_luck=request.POST.get('good_luck') or 0,
                expense_type=request.POST['expense_type'],
                expense_amt=request.POST.get('expense_amt') or 0,
                bill_no=generate_bill_no(),
                photo=photo_file,
                remarks=request.POST['remarks'],
                date=now.date(),
                time=now.time(),
                entry_source='customer_staff_entry'
            )

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'id': entry.id})

        messages.success(request, "✅ Data saved successfully!")
        return redirect('staff_entry')

    context = {
        'now': now,
        'customer_id': "cust-" + str(random.randint(10000, 99999)),
        'bill_no': generate_bill_no(),
        'stores':store,
        'staff_machine':staff_machine,
        'machines':machines,
    }
    return render(request, 'customer_staff_entry.html', context)

# @login_required
# def all_entries(request):
#     if request.user.role == 'staff':
#         if not request.user.store:
#             messages.error(request, "You are not assigned to any store.")
#             return redirect('custom_login')
#         stores = Store.objects.filter(pk=request.user.store.pk)
#     elif request.user.is_superuser or request.user.role == 'superadmin':
#         stores = Store.objects.all()
#     else:
#         messages.error(request, "Unauthorized access.")
#         return redirect('custom_login')

#     entries = GameData.objects.select_related('machine__store', 'staff').order_by('-id')[:100]
#     return render(request, 'entries_list.html', {
#         'entries': entries,
#         'stores': stores
#     })

@login_required
def staff_entries(request):
    if request.user.role == 'staff':
        if not request.user.store:
            messages.error(request, "You are not assigned to any store.")
            return redirect('custom_login')
        
        # Staff should only see their own entries
        entries = GameData.objects.filter(
            entry_source='staff_entry',
            staff=request.user
        ).select_related('machine__store', 'staff').order_by('-id')[:100]

        stores = Store.objects.filter(pk=request.user.store.pk)

    elif request.user.is_superuser or request.user.role == 'superadmin':
        entries = GameData.objects.filter(entry_source='staff_entry') \
                                  .select_related('machine__store', 'staff') \
                                  .order_by('-id')[:100]
        stores = Store.objects.all()
    else:
        messages.error(request, "Unauthorized access.")
        return redirect('custom_login')

    return render(request, 'staff_entries_list.html', {
        'entries': entries,
        'stores': stores
    })

@login_required
def customer_entries(request):
    if request.user.role == 'staff':
        if not request.user.store:
            messages.error(request, "You are not assigned to any store.")
            return redirect('custom_login')

        entries = GameData.objects.filter(
            entry_source='customer_staff_entry',
            staff=request.user
        ).select_related('machine__store', 'staff').order_by('-id')[:100]

        stores = Store.objects.filter(pk=request.user.store.pk)

    elif request.user.is_superuser or request.user.role == 'superadmin':
        entries = GameData.objects.filter(entry_source='customer_staff_entry') \
                                  .select_related('machine__store', 'staff') \
                                  .order_by('-id')[:100]
        stores = Store.objects.all()
    else:
        messages.error(request, "Unauthorized access.")
        return redirect('custom_login')

    return render(request, 'customer_entries_list.html', {
        'entries': entries,
        'stores': stores
    })
